#!/usr/bin/env python3
"""
nbtparsedat v4 - Universal Minecraft Seed Recovery
Extracts seeds from ANY file: corrupt, recovered, wrong extension, fragmented.
Works on Windows, Linux, macOS without editing the script.
"""

import sys, os, re, gzip, zlib, struct, traceback
from pathlib import Path
from collections import defaultdict
from datetime import datetime

# Optional deps
try:
    import nbtlib
    HAS_NBT = True
except ImportError:
    HAS_NBT = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

# ── Config ──
MIN_FILE_SIZE = 50
MAX_FILE_SIZE = 500 * 1024 * 1024
MIN_SEED_LEN = 5
MAX_SEED_LEN = 19
MAX_GZIP_FRAGS = 50
MAX_DECOMP = 32 * 1024 * 1024
KNOWN_SEEDS = {"1234567890", "9876543210"}

# Magic bytes to SKIP (non-Minecraft binary formats)
SKIP_MAGICS = [
    b'\x89PNG', b'\xff\xd8\xff', b'GIF8', b'BM',  # images
    b'\x7fELF', b'MZ',  # executables
    b'%PDF', b'\xd0\xcf\x11\xe0',  # documents
    b'PK\x03\x04',  # zip (not gzip)
    b'Rar!', b'7z\xbc\xaf',  # archives
    b'ID3', b'\xff\xfb', b'fLaC',  # audio
    b'ftyp', b'RIFF',  # video
    b'SQLite',  # databases
]

MINECRAFT_STRONG = [b'RandomSeed', b'worldGenSeed', b'LevelName', b'WorldGenSettings',
                    b'SpawnX', b'SpawnY', b'SpawnZ', b'GameType', b'Player']
MINECRAFT_MEDIUM = [b'seed', b'Seed', b'level.dat', b'LastPlayed', b'Time',
                    b'DayTime', b'Dimension', b'Data', b'minecraft']
ANTI_PATTERNS = [b'javascript', b'function(', b'var ', b'const ', b'window.',
                 b'document.', b'<html', b'<script', b'HTTP/', b'React',
                 b'jQuery', b'module.exports', b'Copyright', b'MIT License']

# Regex patterns for text-based seed extraction
TEXT_SEED_PATTERNS = [
    re.compile(r'RandomSeed[:\s]+(-?\d{5,19})', re.I),
    re.compile(r'worldGenSeed[:\s]+(-?\d{5,19})', re.I),
    re.compile(r'[Ss]eed[:\s"\'=]+(-?\d{5,19})'),
    re.compile(r'Seed:\s*\[(-?\d{5,19})\]'),
    re.compile(r'/seed\s+(-?\d{5,19})'),
    re.compile(r'using\s+seed\s*[:\[]?\s*(-?\d{5,19})', re.I),
    re.compile(r'world\s+seed[:\s]+(-?\d{5,19})', re.I),
    re.compile(r'seed\s*=\s*(-?\d{5,19})', re.I),
    re.compile(r'generating\s+with\s+seed[:\s]+(-?\d{5,19})', re.I),
    re.compile(r'Random\s+seed[:\s]+(-?\d{5,19})', re.I),
    re.compile(r'Seed\s+used[:\s]+(-?\d{5,19})', re.I),
    re.compile(r'(?:world|level|map).*?seed.*?(-?\d{5,19})', re.I),
]
BINARY_SEED_PATTERNS = [re.compile(p.pattern.encode(), p.flags & ~(re.UNICODE)) for p in TEXT_SEED_PATTERNS[:5]]


def sanitize(text):
    if text is None: return ""
    text = str(text)
    text = ''.join(c if (ord(c) >= 32 or c in '\t\n\r') else '' for c in text)
    return text[:32000]


class SeedHunter:
    def __init__(self, scan_dir):
        self.scan_dir = scan_dir
        self.registry = defaultdict(lambda: {
            'seed': '', 'times_found': 0, 'files': [], 'methods': defaultdict(int),
            'best_confidence': 0, 'best_info': {}, 'world_names': set(), 'game_modes': set()
        })
        self.stats = defaultdict(int)
        self.nbt_entries = []
        self.conf_entries = []  # (seed, score, method, filename, filepath, context_str, nbt_info)

    # ── Confidence Scoring ──
    def calc_confidence(self, seed, context_bytes, method):
        score, reasons = 0, []
        try:
            s = str(seed)
            if s in ('0','1','2','3','4','5','10','100','1000','10000'): return 0, ['Common number']
            if len(s) < MIN_SEED_LEN: return 0, ['Too short']
            if s.replace('-','') == '0' * len(s.replace('-','')): return 0, ['All zeros']
            if len(set(s.replace('-',''))) == 1: return 0, ['Repeated digit']
            if 10 <= len(s) <= 19: score += 20; reasons.append(f'Good length ({len(s)})')
            elif 7 <= len(s) <= 9: score += 10
        except: return 0, ['Invalid']

        if 'nbt' in method.lower(): score += 35; reasons.append('NBT parsed')
        elif 'struct' in method.lower(): score += 30; reasons.append('Binary struct')
        elif 'text' in method.lower(): score += 25; reasons.append('Text pattern')
        elif 'frag' in method.lower(): score += 28; reasons.append('Fragment recovery')

        if isinstance(context_bytes, bytes):
            strong = sum(1 for p in MINECRAFT_STRONG if p in context_bytes)
            medium = sum(1 for p in MINECRAFT_MEDIUM if p in context_bytes)
            if strong: score += min(strong * 15, 40); reasons.append(f'{strong} strong indicators')
            if medium: score += min(medium * 5, 15); reasons.append(f'{medium} medium indicators')
            anti = sum(1 for p in ANTI_PATTERNS if p in context_bytes)
            if anti > 3: return 0, ['Anti-patterns']
            if anti: score -= min(anti * 10, 30)
            if b'RandomSeed' in context_bytes or b'worldGenSeed' in context_bytes:
                score += 20; reasons.append('Direct seed field')

        return max(0, min(100, score)), reasons

    # ── Register a found seed ──
    def register(self, seed, filename, filepath, method, nbt_info, context_bytes):
        seed = str(seed).strip()
        if seed in KNOWN_SEEDS: self.stats['known_excluded'] += 1; return
        score, reasons = self.calc_confidence(seed, context_bytes, method)
        if score <= 0: self.stats['rejected'] += 1; return

        reg = self.registry[seed]
        reg['seed'] = seed; reg['times_found'] += 1
        reg['files'].append(filename); reg['methods'][method] += 1
        if score > reg['best_confidence']:
            reg['best_confidence'] = score
            reg['best_info'] = {'filename': filename, 'filepath': filepath,
                                'method': method, 'reasons': reasons, 'nbt_info': nbt_info}
        if nbt_info.get('world_name'): reg['world_names'].add(nbt_info['world_name'])
        if nbt_info.get('game_mode'): reg['game_modes'].add(nbt_info['game_mode'])

        ctx_str = context_bytes[:100].decode('utf-8', errors='ignore') if isinstance(context_bytes, bytes) else ''
        self.conf_entries.append((seed, score, method, filename, filepath, ctx_str, nbt_info))
        self.stats['pattern_matches'] += 1

    # ── Strategy: NBT parse ──
    def try_nbt_parse(self, data, filepath, filename, method_prefix=''):
        info = {}
        parsed = False
        if HAS_NBT:
            try:
                from io import BytesIO
                nbt = nbtlib.load(BytesIO(data))
                d = nbt.root if hasattr(nbt, 'root') else nbt
                dt = d.get('Data', d)
                seed = dt.get('RandomSeed', dt.get('worldGenSeed', None))
                if seed is None:
                    wgs = dt.get('WorldGenSettings', {})
                    if isinstance(wgs, dict): seed = wgs.get('seed')
                if seed is not None:
                    info['seed'] = str(int(seed))
                    info['world_name'] = str(dt.get('LevelName', ''))
                    gt = dt.get('GameType', -1)
                    info['game_mode'] = {0:'Survival',1:'Creative',2:'Adventure',3:'Spectator'}.get(int(gt) if gt is not None else -1, '')
                    v = dt.get('Version', {})
                    info['version'] = str(v.get('Name', '')) if isinstance(v, dict) else str(v)
                    lp = dt.get('LastPlayed', None)
                    if lp:
                        try: info['last_played'] = datetime.fromtimestamp(int(lp)/1000).strftime('%Y-%m-%d %H:%M:%S')
                        except: info['last_played'] = str(lp)
                    info['time_played'] = str(dt.get('Time', ''))
                    sx, sy, sz = dt.get('SpawnX',0), dt.get('SpawnY',0), dt.get('SpawnZ',0)
                    info['spawn'] = f"({sx},{sy},{sz})"
                    info['hardcore'] = str(dt.get('hardcore', ''))
                    info['difficulty'] = str(dt.get('Difficulty', ''))
                    info['generator'] = str(dt.get('generatorName', ''))
                    parsed = True
                    self.stats['nbt_parsed'] += 1
                    m = f'NBT-{method_prefix}' if method_prefix else 'NBT-direct'
                    self.register(info['seed'], filename, filepath, m, info, data[:1000])
                    self.nbt_entries.append((filename, info, filepath))
            except: pass

        # Manual fallback
        if not parsed and (b'RandomSeed' in data or b'worldGenSeed' in data):
            for tag_name in [b'RandomSeed', b'worldGenSeed']:
                idx = data.find(tag_name)
                if idx == -1: continue
                for off in range(idx + len(tag_name), min(idx + len(tag_name) + 50, len(data) - 7)):
                    try:
                        val = struct.unpack('>q', data[off:off+8])[0]
                        if 1000 < abs(val) < 9223372036854775807:
                            info['seed'] = str(val)
                            # Try world name
                            li = data.find(b'LevelName')
                            if li != -1:
                                nd = data[li+10:li+100]
                                for i in range(len(nd)-1):
                                    if nd[i] == 0 and nd[i+1] > 0:
                                        nm = nd[i+1:i+50].split(b'\x00')[0].decode('utf-8', errors='ignore')
                                        if nm and 1 < len(nm) < 50: info['world_name'] = nm; break
                            self.stats['nbt_parsed'] += 1
                            m = f'NBT-manual-{method_prefix}' if method_prefix else 'NBT-manual'
                            self.register(info['seed'], filename, filepath, m, info, data[max(0,idx-200):idx+200])
                            self.nbt_entries.append((filename, info, filepath))
                            return
                    except: pass

    # ── Strategy: Binary struct TAG_Long extraction ──
    def scan_binary_tags(self, data, filepath, filename):
        # TAG_Long (0x04) + 2-byte name len + name + 8-byte big-endian int64
        for tag_name, label in [(b'RandomSeed', 'RandomSeed'), (b'seed', 'WGS-seed')]:
            name_len = len(tag_name)
            expected = b'\x04' + struct.pack('>H', name_len) + tag_name
            pos = 0
            while True:
                idx = data.find(expected, pos)
                if idx == -1: break
                val_start = idx + len(expected)
                if val_start + 8 <= len(data):
                    try:
                        val = struct.unpack('>q', data[val_start:val_start+8])[0]
                        if abs(val) > 1000:
                            ctx = data[max(0,idx-200):min(len(data),val_start+208)]
                            self.register(str(val), filename, filepath, f'Struct-{label}', {}, ctx)
                    except: pass
                pos = idx + 1

    # ── Strategy: Text/log pattern scan ──
    def scan_text(self, data, filepath, filename, method_prefix=''):
        try:
            text = data.decode('utf-8', errors='ignore')
            printable = sum(1 for c in text[:1000] if c.isprintable() or c in '\n\r\t')
            if printable / max(len(text[:1000]), 1) < 0.3: return
        except: return

        self.stats['text_searches'] += 1
        for pat in TEXT_SEED_PATTERNS:
            for m in pat.finditer(text):
                seed = m.group(1)
                start = max(0, m.start()-200)
                ctx = text[start:m.end()+200].encode('utf-8', errors='ignore')
                method = f'Text-{method_prefix}' if method_prefix else 'Text-raw'
                self.register(seed, filename, filepath, method, {}, ctx)

    # ── Strategy: Binary regex patterns ──
    def scan_binary_patterns(self, data, filepath, filename, method_prefix=''):
        for pat in BINARY_SEED_PATTERNS:
            for m in pat.finditer(data):
                try:
                    seed = m.group(1).decode('ascii', errors='ignore')
                    start = max(0, m.start()-200)
                    ctx = data[start:m.end()+200]
                    method = f'Binary-{method_prefix}' if method_prefix else 'Binary-raw'
                    self.register(seed, filename, filepath, method, {}, ctx)
                except: pass

    # ── Strategy: Multi-decompression ──
    def try_decompress(self, data):
        results = [('raw', data)]
        try: results.append(('gzip', gzip.decompress(data))); self.stats['decompressions'] += 1
        except: pass
        for wb in [15, -15, 15+16, 15+32]:
            try: results.append((f'zlib-{wb}', zlib.decompress(data, wb))); self.stats['decompressions'] += 1
            except: pass
        return results

    # ── Strategy: Fragmented gzip recovery ──
    def scan_gzip_fragments(self, data, filepath, filename):
        count = 0
        pos = 0
        while count < MAX_GZIP_FRAGS:
            idx = data.find(b'\x1f\x8b', pos)
            if idx == -1 or idx == 0: break  # skip offset 0, already tried
            pos = idx + 1; count += 1
            try:
                chunk = gzip.decompress(data[idx:idx+MAX_DECOMP])
                if len(chunk) < 10: continue
                self.stats['frag_decompressions'] += 1
                self.try_nbt_parse(chunk, filepath, filename, f'frag@{idx}')
                self.scan_binary_tags(chunk, filepath, filename + f'@frag{idx}')
                self.scan_text(chunk, filepath, filename, f'frag@{idx}')
            except: pass

    # ── Should we skip this file? ──
    def should_skip(self, filepath):
        try:
            sz = os.path.getsize(filepath)
            if sz < MIN_FILE_SIZE or sz > MAX_FILE_SIZE: return True
            with open(filepath, 'rb') as f:
                header = f.read(8)
            for magic in SKIP_MAGICS:
                if header.startswith(magic): return True
            return False
        except: return True

    # ── Process one file ──
    def process_file(self, filepath, filename):
        self.stats['files_scanned'] += 1
        try:
            with open(filepath, 'rb') as f:
                data = f.read(min(MAX_FILE_SIZE, os.path.getsize(filepath)))
        except: return

        # Try all decompression variants
        for method, ddata in self.try_decompress(data):
            self.try_nbt_parse(ddata, filepath, filename, method)
            self.scan_binary_tags(ddata, filepath, filename)
            self.scan_binary_patterns(ddata, filepath, filename, method)
            self.scan_text(ddata, filepath, filename, method)

        # Fragmented gzip recovery
        self.scan_gzip_fragments(data, filepath, filename)

    # ── Collect and process all files ──
    def run(self):
        print(f"\n{'='*60}")
        print(f"  MINECRAFT SEED RECOVERY v4")
        print(f"{'='*60}")
        if not HAS_NBT: print("  [!] nbtlib not installed - using manual parsing only")
        if not HAS_XLSX: print("  [!] openpyxl not installed - will output CSV")
        print(f"  Scanning: {self.scan_dir}")
        print(f"{'='*60}\n")

        # Collect files
        files = []
        print("Collecting files...", end='', flush=True)
        for root, dirs, fnames in os.walk(self.scan_dir):
            for fn in fnames:
                fp = os.path.join(root, fn)
                if not self.should_skip(fp):
                    files.append((fp, fn))
                else:
                    self.stats['files_skipped'] += 1
        print(f" {len(files)} files to scan ({self.stats['files_skipped']} skipped)")

        if not files:
            print("No files found!"); return

        # Process
        last_pct = -1
        for i, (fp, fn) in enumerate(files, 1):
            pct = int(i / len(files) * 100)
            if pct != last_pct:
                bar = '#' * (pct // 2) + '.' * (50 - pct // 2)
                print(f"\r[{bar}] {pct}% ({i}/{len(files)})", end='', flush=True)
                last_pct = pct
            try: self.process_file(fp, fn)
            except: pass

        print(f"\r[{'#'*50}] 100% ({len(files)}/{len(files)})")
        self.save_results()
        self.print_summary()

    # ── Save results ──
    def save_results(self):
        print("\nWriting results...")
        if HAS_XLSX:
            self._save_xlsx()
        else:
            self._save_csv()

    def _save_xlsx(self):
        wb = Workbook()
        # Sheet 1: All Seeds summary
        ws = wb.active; ws.title = "All Seeds"
        headers = ['Seed','Times Found','Best Confidence','Methods','World Names',
                    'Game Modes','First Found In','Path']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(1, c, h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')

        row = 2
        for seed, reg in sorted(self.registry.items(), key=lambda x: (-x[1]['best_confidence'], -x[1]['times_found'])):
            methods = ', '.join(f"{m}({c}x)" for m, c in reg['methods'].items())
            ws.cell(row, 1, sanitize(seed))
            ws.cell(row, 2, reg['times_found'])
            ws.cell(row, 3, reg['best_confidence'])
            ws.cell(row, 4, sanitize(methods))
            ws.cell(row, 5, sanitize(', '.join(reg['world_names'])))
            ws.cell(row, 6, sanitize(', '.join(reg['game_modes'])))
            ws.cell(row, 7, sanitize(reg['files'][0] if reg['files'] else ''))
            ws.cell(row, 8, sanitize(reg['best_info'].get('filepath', '')))
            row += 1

        # Sheet 2: High/Med/Low confidence
        for label, lo, hi, color in [('High Confidence',70,101,'4CAF50'),('Medium Confidence',40,70,'FF9800'),('Low Confidence',1,40,'FFC107')]:
            ws2 = wb.create_sheet(label)
            h2 = ['Seed','Score','Method','File','Context','World Name','Path']
            for c, h in enumerate(h2, 1):
                cell = ws2.cell(1, c, h)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            r = 2
            for seed, score, method, fn, fp, ctx, nbt in self.conf_entries:
                if lo <= score < hi:
                    ws2.cell(r, 1, sanitize(seed)); ws2.cell(r, 2, score)
                    ws2.cell(r, 3, sanitize(method)); ws2.cell(r, 4, sanitize(fn))
                    ws2.cell(r, 5, sanitize(ctx))
                    ws2.cell(r, 6, sanitize(nbt.get('world_name', '')))
                    ws2.cell(r, 7, sanitize(fp)); r += 1

        # Sheet: NBT Parsed
        ws3 = wb.create_sheet("NBT Parsed")
        h3 = ['File','Seed','World Name','Game Mode','Version','Last Played','Time Played','Spawn','Hardcore','Difficulty','Generator','Path']
        for c, h in enumerate(h3, 1):
            cell = ws3.cell(1, c, h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
        for r, (fn, info, fp) in enumerate(self.nbt_entries, 2):
            for c, k in enumerate(['seed','world_name','game_mode','version','last_played','time_played','spawn','hardcore','difficulty','generator'], 2):
                ws3.cell(r, c, sanitize(info.get(k, '')))
            ws3.cell(r, 1, sanitize(fn)); ws3.cell(r, 12, sanitize(fp))

        # Sheet: Statistics
        ws4 = wb.create_sheet("Statistics")
        stats_data = [('Files Scanned', self.stats['files_scanned']),
                      ('Files Skipped', self.stats['files_skipped']),
                      ('NBT Parsed', self.stats['nbt_parsed']),
                      ('Decompressions', self.stats['decompressions']),
                      ('Fragment Decompressions', self.stats['frag_decompressions']),
                      ('Text Searches', self.stats['text_searches']),
                      ('Pattern Matches', self.stats['pattern_matches']),
                      ('Known Seeds Excluded', self.stats['known_excluded']),
                      ('Rejected (low quality)', self.stats['rejected']),
                      ('Unique Seeds Found', len(self.registry))]
        ws4.cell(1, 1, 'Metric'); ws4.cell(1, 2, 'Value')
        for r, (k, v) in enumerate(stats_data, 2):
            ws4.cell(r, 1, k); ws4.cell(r, 2, v)

        # Auto-width
        for ws in wb.worksheets:
            for col in ws.columns:
                mx = max((len(str(c.value or '')) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(mx + 2, 80)

        out = os.path.join(self.scan_dir, "minecraft_seed_recovery_v4.xlsx")
        try:
            wb.save(out); print(f"Saved: {out}")
        except Exception as e:
            print(f"Save failed ({e}), trying next to script...")
            out2 = os.path.join(os.path.dirname(os.path.abspath(__file__)), "minecraft_seed_recovery_v4.xlsx")
            try: wb.save(out2); print(f"Saved: {out2}")
            except Exception as e2: print(f"Save failed: {e2}"); self._save_csv()

    def _save_csv(self):
        import csv
        out = os.path.join(self.scan_dir, "minecraft_seed_recovery_v4.csv")
        try:
            with open(out, 'w', newline='', encoding='utf-8') as f:
                w = csv.writer(f)
                w.writerow(['Seed','Times Found','Confidence','Methods','World Names','Game Modes','First File','Path'])
                for seed, reg in sorted(self.registry.items(), key=lambda x: -x[1]['best_confidence']):
                    methods = ', '.join(f"{m}({c}x)" for m, c in reg['methods'].items())
                    w.writerow([seed, reg['times_found'], reg['best_confidence'], methods,
                                ', '.join(reg['world_names']), ', '.join(reg['game_modes']),
                                reg['files'][0] if reg['files'] else '', reg['best_info'].get('filepath','')])
            print(f"Saved CSV: {out}")
        except Exception as e:
            print(f"CSV save failed: {e}")

    def print_summary(self):
        high = sum(1 for r in self.registry.values() if r['best_confidence'] >= 70)
        med = sum(1 for r in self.registry.values() if 40 <= r['best_confidence'] < 70)
        low = sum(1 for r in self.registry.values() if r['best_confidence'] < 40)
        print(f"\n{'='*60}")
        print(f"  SCAN COMPLETE")
        print(f"{'='*60}")
        print(f"  Files Scanned:        {self.stats['files_scanned']}")
        print(f"  Files Skipped:        {self.stats['files_skipped']}")
        print(f"  NBT Parsed:           {self.stats['nbt_parsed']}")
        print(f"  Decompressions:       {self.stats['decompressions']}")
        print(f"  Fragment Recoveries:  {self.stats['frag_decompressions']}")
        print(f"  Text Searches:        {self.stats['text_searches']}")
        print(f"  Seeds Rejected:       {self.stats['rejected']}")
        print(f"  ─────────────────────────────")
        print(f"  UNIQUE SEEDS FOUND:   {len(self.registry)}")
        print(f"    High Confidence:    {high}")
        print(f"    Medium Confidence:  {med}")
        print(f"    Low Confidence:     {low}")
        print(f"{'='*60}")
        if self.registry:
            print("\n  Check 'All Seeds' sheet first!")
            print("  High confidence seeds are most likely real.")


def main():
    print("\n" + "="*60)
    print("  MINECRAFT SEED RECOVERY v4")
    print("  Works with corrupt, recovered, and misidentified files")
    print("="*60)

    # Interactive directory prompt
    if len(sys.argv) > 1:
        scan_dir = sys.argv[1]
    else:
        default = os.getcwd()
        print(f"\n  Default directory: {default}")
        user_input = input("\n  Enter directory to scan (or press Enter for default): ").strip()
        scan_dir = user_input if user_input else default

    scan_dir = os.path.normpath(os.path.expanduser(scan_dir))
    if not os.path.isdir(scan_dir):
        print(f"\n  ERROR: Directory not found: {scan_dir}")
        input("\n  Press Enter to exit...")
        return

    hunter = SeedHunter(scan_dir)
    hunter.run()
    input("\n  Press Enter to exit...")


if __name__ == '__main__':
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nInterrupted by user")
    except Exception as e:
        print(f"\nFatal error: {e}")
        traceback.print_exc()
        input("\nPress Enter to exit...")
